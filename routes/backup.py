import io
import zipfile
import unicodedata
from datetime import date
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db
from models.client import Cliente
from models.prestamo import Prestamo
from models.cuota import Cuota
from models.pago import Pago
from models.archivo import Archivo
from services.auth import get_current_user

router = APIRouter()

ACCENT = "0284C7"
GREEN  = "16A34A"
ORANGE = "EA580C"
PURPLE = "7C3AED"


def _header_row(ws, headers: list[str], fill_color: str):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autowidth(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def _safe(texto: str) -> str:
    """Elimina tildes y deja solo caracteres seguros para nombres de archivo."""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return "".join(c for c in texto if c.isalnum() or c in ("_", "-")).strip()


@router.get("/zip")
def backup_zip(
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """
    Genera un ZIP con el backup completo:
    - backup_YYYY-MM-DD.xlsx  (Resumen, Clientes, Préstamos, Cuotas, Pagos)
    - archivos/<Apellido_Nombre_DNI>/pagare.pdf           (si existe)
    - archivos/<Apellido_Nombre_DNI>/recibo_sueldo.pdf    (si existe)
    """
    today = date.today().isoformat()

    # ── Cargar datos ─────────────────────────────────────────────────────────
    clientes = db.query(Cliente).order_by(Cliente.apellido, Cliente.nombre).all()
    prestamos = (
        db.query(Prestamo, Cliente)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .order_by(Prestamo.id)
        .all()
    )
    cuotas = (
        db.query(Cuota, Prestamo, Cliente)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .order_by(Cuota.prestamo_id, Cuota.numero_cuota)
        .all()
    )
    pagos = (
        db.query(Pago, Prestamo, Cliente)
        .join(Prestamo, Pago.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .order_by(Pago.fecha_pago.desc())
        .all()
    )
    archivos = db.query(Archivo, Cliente).join(Cliente, Archivo.cliente_id == Cliente.id).all()

    # ── Construir Excel ───────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Hoja: Clientes
    ws_c = wb.create_sheet("Clientes")
    _header_row(ws_c, ["ID", "Apellido", "Nombre", "DNI", "Teléfono", "Domicilio", "Empleo", "Pagaré", "Recibo Sueldo", "Cliente Desde"], ACCENT)
    archivos_por_cliente: dict[int, set[str]] = {}
    for a, _ in archivos:
        archivos_por_cliente.setdefault(a.cliente_id, set()).add(a.tipo)
    for c in clientes:
        tipos = archivos_por_cliente.get(c.id, set())
        ws_c.append([
            c.id,
            c.apellido,
            c.nombre,
            c.dni,
            c.telefono or "",
            c.domicilio or "",
            c.empleo or "",
            "Sí" if "pagare" in tipos else "",
            "Sí" if "recibo_sueldo" in tipos else "",
            c.fecha_creacion.strftime("%Y-%m-%d") if c.fecha_creacion else "",
        ])
    _autowidth(ws_c)

    # Hoja: Préstamos
    ws_p = wb.create_sheet("Préstamos")
    _header_row(ws_p, [
        "ID", "Cliente ID", "Apellido", "Nombre", "DNI",
        "Monto", "Interés %", "Cuotas", "Tipo", "Fecha Inicio", "Estado", "Notas"
    ], GREEN)
    for p, c in prestamos:
        ws_p.append([
            p.id, c.id, c.apellido, c.nombre, c.dni,
            float(p.monto), p.interes_total, p.cuotas,
            p.tipo_prestamo or "mensual",
            p.fecha_inicio.isoformat() if p.fecha_inicio else "",
            p.estado, p.notas or "",
        ])
    _autowidth(ws_p)

    # Hoja: Cuotas
    ws_q = wb.create_sheet("Cuotas")
    _header_row(ws_q, [
        "ID", "Préstamo ID", "Cliente ID", "Apellido", "Nombre",
        "N° Cuota", "Vencimiento", "Monto", "Estado"
    ], ORANGE)
    for c, p, cl in cuotas:
        ws_q.append([
            c.id, p.id, cl.id, cl.apellido, cl.nombre,
            c.numero_cuota,
            c.fecha_vencimiento.isoformat() if c.fecha_vencimiento else "",
            float(c.monto), c.estado,
        ])
    _autowidth(ws_q)

    # Hoja: Pagos
    ws_pg = wb.create_sheet("Pagos")
    _header_row(ws_pg, [
        "ID", "Préstamo ID", "Cliente ID", "Apellido", "Nombre",
        "Monto Pagado", "Fecha Pago", "Días Atraso"
    ], PURPLE)
    for pg, p, cl in pagos:
        ws_pg.append([
            pg.id, p.id, cl.id, cl.apellido, cl.nombre,
            float(pg.monto_pagado),
            pg.fecha_pago.strftime("%Y-%m-%d %H:%M") if pg.fecha_pago else "",
            pg.dias_atraso or 0,
        ])
    _autowidth(ws_pg)

    # Hoja: Resumen (primera)
    ws_r = wb.create_sheet("Resumen", 0)
    _header_row(ws_r, ["Dato", "Valor"], ACCENT)
    total_prestado = db.query(func.coalesce(func.sum(Prestamo.monto), 0)).scalar()
    total_cobrado  = db.query(func.coalesce(func.sum(Pago.monto_pagado), 0)).scalar()
    ws_r.append(["Fecha del backup", today])
    ws_r.append(["Total clientes", len(clientes)])
    ws_r.append(["Total préstamos", len(prestamos)])
    ws_r.append(["Préstamos activos", sum(1 for p, _ in prestamos if p.estado == "activo")])
    ws_r.append(["Total prestado", float(total_prestado)])
    ws_r.append(["Total cobrado", float(total_cobrado)])
    ws_r.append(["Total cuotas", len(cuotas)])
    ws_r.append(["Total pagos", len(pagos)])
    ws_r.append(["Archivos PDF", len(archivos)])
    _autowidth(ws_r)

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_bytes = excel_buf.getvalue()

    # ── Construir ZIP ─────────────────────────────────────────────────────────
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        # Excel principal
        zf.writestr(f"backup_{today}.xlsx", excel_bytes)

        # PDFs organizados por cliente
        for a, c in archivos:
            carpeta = f"{_safe(c.apellido)}_{_safe(c.nombre)}_{_safe(c.dni)}"
            nombre_pdf = f"{'pagare' if a.tipo == 'pagare' else 'recibo_sueldo'}.pdf"
            zf.writestr(f"archivos/{carpeta}/{nombre_pdf}", bytes(a.contenido))

    zip_buf.seek(0)
    filename = f"backup_{today}.zip"
    return StreamingResponse(
        iter([zip_buf.read()]),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
