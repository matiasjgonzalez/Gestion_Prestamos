import io
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models.archivo import Archivo
from services.mora_service import verificar_mora, obtener_cuotas_en_mora
from services.auth import get_current_user

router = APIRouter()


@router.post("/verificar")
def verificar(
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    nuevas_vencidas = verificar_mora(db)
    return {
        "nuevas_cuotas_vencidas": len(nuevas_vencidas),
        "detalle": nuevas_vencidas,
    }


@router.get("/")
def listar_mora(
    search: str = "",
    limit: int = 10,
    offset: int = 0,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    result = obtener_cuotas_en_mora(db, search=search, limit=limit, offset=offset)
    return {
        "total_en_mora": result["total"],
        "total_monto_mora": result["total_monto"],
        "cuotas": result["cuotas"],
    }


@router.get("/export/zip")
def export_mora_zip(
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Exporta mora como ZIP: mora.xlsx (agrupado por cliente) + archivos PDF adjuntos."""
    result = obtener_cuotas_en_mora(db, limit=100_000)
    cuotas = result["cuotas"]

    # ── Agrupar por cliente ──────────────────────────────────────────────────
    clientes_dict: dict[int, dict] = {}
    for c in cuotas:
        cid = c.get("cliente_id")
        if cid is None:
            continue
        if cid not in clientes_dict:
            clientes_dict[cid] = {
                "nombre": c["cliente_nombre"],
                "dni": c["cliente_dni"],
                "cuotas": 0,
                "monto_total": 0.0,
                "dias_max_atraso": 0,
            }
        entry = clientes_dict[cid]
        entry["cuotas"] += 1
        entry["monto_total"] += c["monto"]
        entry["dias_max_atraso"] = max(entry["dias_max_atraso"], c.get("dias_atraso", 0))

    # ── Construir mora.xlsx ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mora"

    headers = ["Cliente", "DNI", "Cuotas en Mora", "Monto Total ($)", "Días Máx. Atraso", "Tiene Pagare", "Tiene Recibo"]
    header_fill = PatternFill("solid", fgColor="E11D48")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Consulta de archivos para los clientes en mora
    cliente_ids = list(clientes_dict.keys())
    archivos_rows = (
        db.query(Archivo.cliente_id, Archivo.tipo, Archivo.nombre_archivo, Archivo.contenido)
        .filter(Archivo.cliente_id.in_(cliente_ids))
        .all()
    ) if cliente_ids else []

    # Indexar por cliente_id → {tipo: (nombre, contenido)}
    archivos_map: dict[int, dict] = {}
    for a in archivos_rows:
        if a.cliente_id not in archivos_map:
            archivos_map[a.cliente_id] = {}
        archivos_map[a.cliente_id][a.tipo] = (a.nombre_archivo, a.contenido)

    for cid, entry in sorted(clientes_dict.items(), key=lambda x: x[1]["nombre"] or ""):
        tiene_pagare = "Si" if archivos_map.get(cid, {}).get("pagare") else "No"
        tiene_recibo = "Si" if archivos_map.get(cid, {}).get("recibo_sueldo") else "No"
        ws.append([
            entry["nombre"],
            entry["dni"],
            entry["cuotas"],
            round(entry["monto_total"], 2),
            entry["dias_max_atraso"],
            tiene_pagare,
            tiene_recibo,
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(cell.value or "")) for cell in col
        ) + 4

    xlsx_bytes = io.BytesIO()
    wb.save(xlsx_bytes)
    xlsx_bytes.seek(0)

    # ── Construir ZIP ────────────────────────────────────────────────────────
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("mora.xlsx", xlsx_bytes.read())
        for cid, files in archivos_map.items():
            for tipo, (nombre_archivo, contenido) in files.items():
                zf.writestr(f"archivos/{nombre_archivo}", contenido)

    zip_buffer.seek(0)
    return StreamingResponse(
        iter([zip_buffer.read()]),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=mora.zip"},
    )
