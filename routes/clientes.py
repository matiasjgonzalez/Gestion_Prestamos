import io
import unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, or_, and_
from datetime import date
from database import get_db
from models.client import Cliente
from models.prestamo import Prestamo
from models.cuota import Cuota
from models.pago import Pago
from models.archivo import Archivo
from schemas.client import ClienteCreate, ClienteRead, ClienteUpdate
from services.auth import get_current_user

TIPOS_ARCHIVO = {"pagare", "recibo_sueldo"}
MAX_ARCHIVO_BYTES = 2 * 1024 * 1024  # 2 MB


def _normalizar(texto: str) -> str:
    """Elimina tildes y caracteres especiales, devuelve solo ASCII alfanumérico."""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return "".join(c for c in texto if c.isalnum() or c in (" ", "_", "-")).strip()

router = APIRouter()


@router.post("/", response_model=ClienteRead)
def create_cliente(
    payload: ClienteCreate,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    existente = db.query(Cliente).filter(Cliente.dni == payload.dni).first()
    if existente:
        raise HTTPException(status_code=400, detail="DNI ya registrado")
    cliente = Cliente(**payload.model_dump())
    db.add(cliente)
    db.commit()
    db.refresh(cliente)
    return cliente


@router.get("/export/zip")
def export_clientes_zip(
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """
    Exporta todos los clientes en un ZIP:
    - clientes.xlsx  (listado con resumen de préstamos y deuda)
    - archivos/<Apellido_Nombre_DNI>/Pagare_xxx.pdf        (si existe)
    - archivos/<Apellido_Nombre_DNI>/ReciboSueldo_xxx.pdf  (si existe)
    """
    import zipfile as zf_mod
    from sqlalchemy import func as sf

    clientes = db.query(Cliente).order_by(Cliente.apellido, Cliente.nombre).all()

    # Total préstamos por cliente (dos queries simples, sin func.case)
    total_rows = (
        db.query(Prestamo.cliente_id, sf.count(Prestamo.id).label("total"))
        .group_by(Prestamo.cliente_id)
        .all()
    )
    activos_rows = (
        db.query(Prestamo.cliente_id, sf.count(Prestamo.id).label("activos"))
        .filter(Prestamo.estado == "activo")
        .group_by(Prestamo.cliente_id)
        .all()
    )
    totales  = {r.cliente_id: int(r.total)   for r in total_rows}
    activos  = {r.cliente_id: int(r.activos) for r in activos_rows}

    # Deuda pendiente por cliente
    deuda_rows = (
        db.query(Prestamo.cliente_id, sf.coalesce(sf.sum(Cuota.monto), 0).label("deuda"))
        .join(Cuota, Cuota.prestamo_id == Prestamo.id)
        .filter(Cuota.estado.in_(["pendiente", "vencida"]))
        .group_by(Prestamo.cliente_id)
        .all()
    )
    deuda = {r.cliente_id: float(r.deuda) for r in deuda_rows}

    # Archivos PDF agrupados por cliente
    archivos_rows = db.query(Archivo).all()
    archivos_por_cliente: dict[int, list[Archivo]] = {}
    for a in archivos_rows:
        archivos_por_cliente.setdefault(a.cliente_id, []).append(a)

    # ── Construir Excel ───────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = ["ID", "Apellido", "Nombre", "DNI", "Teléfono", "Domicilio", "Empleo",
               "Préstamos Totales", "Préstamos Activos", "Deuda Pendiente",
               "Pagaré", "Recibo Sueldo", "Cliente desde"]
    header_fill = PatternFill("solid", fgColor="0284C7")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for c in clientes:
        tipos = {a.tipo for a in archivos_por_cliente.get(c.id, [])}
        ws.append([
            c.id,
            c.apellido,
            c.nombre,
            c.dni,
            c.telefono or "",
            c.domicilio or "",
            c.empleo or "",
            totales.get(c.id, 0),
            activos.get(c.id, 0),
            deuda.get(c.id, 0.0),
            "Sí" if "pagare" in tipos else "",
            "Sí" if "recibo_sueldo" in tipos else "",
            c.fecha_creacion.strftime("%Y-%m-%d") if c.fecha_creacion else "",
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in col) + 4, 50
        )

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_bytes = excel_buf.getvalue()

    # ── Construir ZIP ─────────────────────────────────────────────────────────
    zip_buf = io.BytesIO()
    with zf_mod.ZipFile(zip_buf, "w", compression=zf_mod.ZIP_DEFLATED) as zf:
        zf.writestr("clientes.xlsx", excel_bytes)
        for c in clientes:
            archivos = archivos_por_cliente.get(c.id, [])
            if not archivos:
                continue
            carpeta = f"{_normalizar(c.apellido).replace(' ', '')}_{_normalizar(c.nombre).replace(' ', '')}_{_normalizar(c.dni)}"
            for a in archivos:
                zf.writestr(f"archivos/{carpeta}/{a.nombre_archivo}", bytes(a.contenido))

    zip_buf.seek(0)
    return StreamingResponse(
        iter([zip_buf.read()]),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=clientes.zip"},
    )


@router.get("/", response_model=list[ClienteRead])
def list_clientes(
    offset: int = 0,
    limit: int = 100,
    search: str = Query(None, description="Buscar por nombre, apellido o DNI"),
    sort_desc: bool = Query(False, description="Orden descendente por apellido"),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    query = db.query(Cliente)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            (Cliente.nombre.ilike(pattern))
            | (Cliente.apellido.ilike(pattern))
            | (Cliente.dni.ilike(pattern))
            | (func.concat(Cliente.nombre, ' ', Cliente.apellido).ilike(pattern))
        )
    if sort_desc:
        query = query.order_by(Cliente.apellido.desc(), Cliente.nombre.desc())
    else:
        query = query.order_by(Cliente.apellido, Cliente.nombre)
    clientes = query.offset(offset).limit(limit).all()

    # IDs de clientes con mora: ya marcadas como "vencida" O pendientes con fecha pasada
    ids_con_mora = set(
        row[0] for row in db.query(Prestamo.cliente_id)
        .join(Cuota, Cuota.prestamo_id == Prestamo.id)
        .filter(
            or_(
                Cuota.estado == "vencida",
                and_(Cuota.estado == "pendiente", Cuota.fecha_vencimiento < date.today())
            )
        )
        .distinct()
        .all()
    )

    # IDs de clientes con documentos
    ids_con_docs = set(
        row[0] for row in db.query(Archivo.cliente_id).distinct().all()
    )

    result = []
    for c in clientes:
        item = ClienteRead.model_validate(c)
        item.tiene_mora = c.id in ids_con_mora
        item.tiene_documentos = c.id in ids_con_docs
        result.append(item)
    return result


@router.get("/{cliente_id}/resumen")
def get_cliente_resumen(
    cliente_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    prestamos = db.query(Prestamo).filter(Prestamo.cliente_id == cliente_id).all()
    prestamo_ids = [p.id for p in prestamos]

    total_cuotas = db.query(func.sum(Cuota.monto)).filter(Cuota.prestamo_id.in_(prestamo_ids)).scalar() or 0
    total_pagado = db.query(func.sum(Pago.monto_pagado)).filter(Pago.prestamo_id.in_(prestamo_ids)).scalar() or 0
    monto_mora = db.query(func.sum(Cuota.monto)).filter(
        Cuota.prestamo_id.in_(prestamo_ids),
        or_(
            Cuota.estado == "vencida",
            and_(Cuota.estado == "pendiente", Cuota.fecha_vencimiento < date.today())
        )
    ).scalar() or 0

    return {
        "prestamos_total": len(prestamos),
        "prestamos_activos": sum(1 for p in prestamos if p.estado == "activo"),
        "deuda_total": round(float(total_cuotas) - float(total_pagado), 2),
        "tiene_mora": monto_mora > 0,
        "monto_mora": float(monto_mora),
    }


@router.get("/{cliente_id}", response_model=ClienteRead)
def get_cliente(
    cliente_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return cliente


@router.put("/{cliente_id}", response_model=ClienteRead)
def update_cliente(
    cliente_id: int,
    payload: ClienteUpdate,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(cliente, k, v)
    db.add(cliente)
    db.commit()
    db.refresh(cliente)
    return cliente


@router.get("/{cliente_id}/estado-cuenta/xlsx")
def estado_cuenta_xlsx(
    cliente_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Exporta estado de cuenta completo del cliente en Excel."""
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    prestamos = (
        db.query(Prestamo)
        .options(joinedload(Prestamo.cuotas_rel), joinedload(Prestamo.pagos))
        .filter(Prestamo.cliente_id == cliente_id)
        .order_by(Prestamo.id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"

    accent = "0284C7"
    header_fill = PatternFill("solid", fgColor=accent)
    header_font = Font(bold=True, color="FFFFFF")
    sub_fill    = PatternFill("solid", fgColor="E2E8F0")
    bold        = Font(bold=True)

    def hrow(row, values, fill=None, font=None):
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=v)
            if fill: c.fill = fill
            if font: c.font = font
            c.alignment = Alignment(horizontal="left", vertical="center")

    r = 1
    ws.cell(r, 1, f"Estado de Cuenta — {cliente.nombre} {cliente.apellido}").font = Font(bold=True, size=13)
    r += 1
    ws.cell(r, 1, f"DNI: {cliente.dni}   Teléfono: {cliente.telefono or '—'}   Domicilio: {cliente.domicilio or '—'}")
    r += 2

    for p in prestamos:
        hrow(r, [f"Préstamo #{p.id}", f"Tipo: {p.tipo_prestamo or 'mensual'}", f"Monto: ${float(p.monto):,.2f}",
                 f"Interés: {p.interes_total}%", f"Estado: {p.estado}"], fill=header_fill, font=header_font)
        r += 1

        hrow(r, ["#", "Vencimiento", "Monto", "Estado"], fill=sub_fill, font=bold)
        r += 1
        for c in sorted(p.cuotas_rel, key=lambda x: x.numero_cuota):
            ws.append([""] * 0)
            ws.cell(r, 1, c.numero_cuota)
            ws.cell(r, 2, c.fecha_vencimiento.isoformat())
            ws.cell(r, 3, float(c.monto))
            ws.cell(r, 4, c.estado)
            r += 1

        if p.pagos:
            r += 1
            hrow(r, ["Pagos registrados", "Fecha", "Monto", "Días atraso"], fill=sub_fill, font=bold)
            r += 1
            for pg in sorted(p.pagos, key=lambda x: x.fecha_pago or date.min):
                ws.cell(r, 1, "")
                ws.cell(r, 2, pg.fecha_pago.strftime("%Y-%m-%d") if pg.fecha_pago else "—")
                ws.cell(r, 3, float(pg.monto_pagado))
                ws.cell(r, 4, pg.dias_atraso or 0)
                r += 1
        r += 2

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            (len(str(c.value or "")) for c in col), default=10
        ) + 4

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    nombre = f"{cliente.apellido}_{cliente.nombre}_estado_cuenta.xlsx".replace(" ", "_")
    return StreamingResponse(
        iter([out.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )


@router.get("/import/template")
def download_import_template(
    _user=Depends(get_current_user),
):
    """Descarga una plantilla Excel para importación masiva de clientes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = ["Nombre (*)", "Apellido (*)", "DNI (*)", "Teléfono", "Domicilio", "Empleo"]
    required_fill = PatternFill("solid", fgColor="0284C7")
    optional_fill = PatternFill("solid", fgColor="334155")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = required_fill if "(*)" in h else optional_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Fila de ejemplo
    ws.append(["Juan", "Pérez", "12345678", "351-1234567", "Av. Colón 123", "Comerciante"])

    # Nota al pie
    ws.cell(row=3, column=1, value="(*) Campos obligatorios. El DNI debe ser único.")
    ws.cell(row=3, column=1).font = Font(italic=True, color="94A3B8", size=9)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col
        ) + 6

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        iter([out.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_clientes.xlsx"},
    )


@router.post("/import/xlsx")
async def import_clientes_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Importa clientes desde un archivo Excel. Devuelve resumen de creados / saltados / errores."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx / .xls)")

    contenido = await file.read()
    if len(contenido) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo no puede superar 5 MB")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel")

    # Buscar fila de encabezado (primera fila no vacía)
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="El archivo está vacío o solo tiene encabezado")

    # Detectar columnas por nombre de encabezado (flexible, sin importar el orden)
    header = [str(c).strip().lower().replace("(*)", "").strip() if c else "" for c in rows[0]]
    col_map = {}
    aliases = {
        "nombre": ["nombre"],
        "apellido": ["apellido"],
        "dni": ["dni"],
        "telefono": ["teléfono", "telefono", "tel", "celular"],
        "domicilio": ["domicilio", "dirección", "direccion"],
        "empleo": ["empleo", "ocupación", "ocupacion", "trabajo"],
    }
    for field, possible in aliases.items():
        for i, h in enumerate(header):
            if h in possible:
                col_map[field] = i
                break

    missing = [f for f in ["nombre", "apellido", "dni"] if f not in col_map]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"No se encontraron las columnas requeridas: {', '.join(missing)}. "
                   f"Asegurate de usar la plantilla descargada."
        )

    # Pre-cargar DNIs existentes para validación rápida
    dni_existentes = set(r[0] for r in db.query(Cliente.dni).all())

    creados, saltados, errores = 0, 0, []

    for fila_num, row in enumerate(rows[1:], start=2):
        # Saltar filas de nota/vacías
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        if str(row[col_map["nombre"]] or "").strip().startswith("("):
            continue

        def get(field):
            idx = col_map.get(field)
            val = row[idx] if idx is not None and idx < len(row) else None
            return str(val).strip() if val is not None else ""

        nombre = get("nombre")
        apellido = get("apellido")
        dni = get("dni")

        if not nombre or not apellido or not dni:
            errores.append(f"Fila {fila_num}: Nombre, Apellido y DNI son obligatorios")
            continue

        if dni in dni_existentes:
            saltados += 1
            continue

        try:
            nuevo = Cliente(
                nombre=nombre,
                apellido=apellido,
                dni=dni,
                telefono=get("telefono") or None,
                domicilio=get("domicilio") or None,
                empleo=get("empleo") or None,
            )
            db.add(nuevo)
            db.flush()
            dni_existentes.add(dni)
            creados += 1
        except Exception as e:
            errores.append(f"Fila {fila_num}: {str(e)}")

    if creados > 0:
        db.commit()
    else:
        db.rollback()

    return {
        "creados": creados,
        "saltados_dni_duplicado": saltados,
        "errores": errores,
        "total_procesadas": creados + saltados + len(errores),
    }


@router.delete("/{cliente_id}")
def delete_cliente(
    cliente_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    db.delete(cliente)
    db.commit()
    return {"ok": True}


# ── Archivos ──────────────────────────────────────────────────────────────────

@router.get("/{cliente_id}/archivos")
def list_archivos(
    cliente_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Lista los archivos del cliente (sin el contenido binario)."""
    db.query(Cliente).filter(Cliente.id == cliente_id).first() or _404("Cliente")
    archivos = (
        db.query(Archivo.id, Archivo.tipo, Archivo.nombre_archivo, Archivo.fecha_subida)
        .filter(Archivo.cliente_id == cliente_id)
        .all()
    )
    return [
        {"id": a.id, "tipo": a.tipo, "nombre_archivo": a.nombre_archivo, "fecha_subida": a.fecha_subida}
        for a in archivos
    ]


@router.post("/{cliente_id}/archivos/{tipo}", status_code=201)
async def subir_archivo(
    cliente_id: int,
    tipo: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Sube (o reemplaza) un archivo PDF para el cliente."""
    if tipo not in TIPOS_ARCHIVO:
        raise HTTPException(status_code=400, detail=f"Tipo inválido. Debe ser: {', '.join(TIPOS_ARCHIVO)}")
    if not file.content_type or "pdf" not in file.content_type.lower():
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos PDF")

    contenido = await file.read()
    if len(contenido) > MAX_ARCHIVO_BYTES:
        raise HTTPException(status_code=400, detail="El archivo supera los 2 MB")
    # Validar magic bytes PDF
    if not contenido.startswith(b"%PDF"):
        raise HTTPException(status_code=400, detail="El archivo no es un PDF válido")

    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    # Generar nombre de archivo
    prefijo = "Pagare" if tipo == "pagare" else "ReciboSueldo"
    apellido = _normalizar(cliente.apellido).replace(" ", "")
    nombre = _normalizar(cliente.nombre).replace(" ", "")
    dni = _normalizar(cliente.dni)
    nombre_archivo = f"{prefijo}_{apellido}{nombre}_{dni}.pdf"

    # Reemplazar si ya existe
    existente = db.query(Archivo).filter(Archivo.cliente_id == cliente_id, Archivo.tipo == tipo).first()
    if existente:
        existente.contenido = contenido
        existente.nombre_archivo = nombre_archivo
        db.commit()
        return {"ok": True, "nombre_archivo": nombre_archivo, "reemplazado": True}

    nuevo = Archivo(
        cliente_id=cliente_id,
        tipo=tipo,
        nombre_archivo=nombre_archivo,
        contenido=contenido,
    )
    db.add(nuevo)
    db.commit()
    return {"ok": True, "nombre_archivo": nombre_archivo, "reemplazado": False}


@router.get("/{cliente_id}/archivos/{tipo}/download")
def descargar_archivo(
    cliente_id: int,
    tipo: str,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Descarga el archivo PDF del tipo indicado."""
    archivo = (
        db.query(Archivo)
        .filter(Archivo.cliente_id == cliente_id, Archivo.tipo == tipo)
        .first()
    )
    if not archivo:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return Response(
        content=archivo.contenido,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{archivo.nombre_archivo}"'},
    )


@router.delete("/{cliente_id}/archivos/{tipo}", status_code=200)
def eliminar_archivo(
    cliente_id: int,
    tipo: str,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Elimina el archivo del tipo indicado para el cliente."""
    archivo = (
        db.query(Archivo)
        .filter(Archivo.cliente_id == cliente_id, Archivo.tipo == tipo)
        .first()
    )
    if not archivo:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    db.delete(archivo)
    db.commit()
    return {"ok": True}


def _404(entidad: str):
    raise HTTPException(status_code=404, detail=f"{entidad} no encontrado")
