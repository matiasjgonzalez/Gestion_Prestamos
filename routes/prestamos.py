import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func as sqlfunc, cast, extract, Date as SADate, case, or_, String
from database import get_db
from schemas.prestamo import PrestamoCreate, PrestamoRead
from schemas.cuota import CuotaRead, CuotaUpdate, CuotaInput
from models.prestamo import Prestamo
from models.cuota import Cuota
from models.pago import Pago
from models.client import Cliente
from services.prestamo_service import create_prestamo, calcular_deuda_restante
from services.mora_service import obtener_cuotas_en_mora
from services.auth import get_current_user
from datetime import date, datetime, timezone, timedelta

router = APIRouter()


@router.get("/dashboard")
def dashboard(
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    # ── Querysets con filtros opcionales ──
    qp  = db.query(Prestamo)
    qpg = db.query(Pago)
    qc  = db.query(Cuota)

    if fecha_desde:
        qp  = qp.filter(Prestamo.fecha_inicio >= fecha_desde)
        qpg = qpg.filter(cast(Pago.fecha_pago, SADate) >= fecha_desde)
        qc  = qc.filter(Cuota.fecha_vencimiento >= fecha_desde)
    if fecha_hasta:
        qp  = qp.filter(Prestamo.fecha_inicio <= fecha_hasta)
        qpg = qpg.filter(cast(Pago.fecha_pago, SADate) <= fecha_hasta)
        qc  = qc.filter(Cuota.fecha_vencimiento <= fecha_hasta)

    total_prestado = float(qp.with_entities(sqlfunc.coalesce(sqlfunc.sum(Prestamo.monto), 0)).scalar())
    total_cobrado  = float(qpg.with_entities(sqlfunc.coalesce(sqlfunc.sum(Pago.monto_pagado), 0)).scalar())
    deuda_total    = round(float(
        qc.filter(Cuota.estado.in_(["pendiente", "vencida", "parcial"]))
          .with_entities(sqlfunc.coalesce(sqlfunc.sum(
              case(
                  (Cuota.estado == "parcial", Cuota.monto - sqlfunc.coalesce(Cuota.monto_pagado_parcial, 0)),
                  else_=Cuota.monto,
              )
          ), 0)).scalar()
    ), 2)
    prestamos_activos = qp.filter(Prestamo.estado == "activo").count()
    clientes_count    = qp.with_entities(sqlfunc.count(sqlfunc.distinct(Prestamo.cliente_id))).scalar() or 0

    tipos_rows = qp.with_entities(Prestamo.tipo_prestamo, sqlfunc.count(Prestamo.id)).group_by(Prestamo.tipo_prestamo).all()
    prestamos_por_tipo = [{"tipo": t or "mensual", "cantidad": c} for t, c in tipos_rows]

    estados_rows = qc.with_entities(Cuota.estado, sqlfunc.count(Cuota.id)).group_by(Cuota.estado).all()
    cuotas_por_estado = [{"estado": e, "cantidad": c} for e, c in estados_rows]

    # ── Cobros de hoy y mañana (sin filtro de período) ──
    hoy_d = date.today()

    def _cuotas_dia(dia):
        rows = (
            db.query(Cuota, Prestamo, Cliente)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .filter(
                Cuota.fecha_vencimiento == dia,
                Cuota.estado.in_(["pendiente", "vencida"]),
                Prestamo.estado == "activo",
            )
            .order_by(Cliente.apellido)
            .all()
        )
        return [{"cuota_id": c.id, "prestamo_id": p.id, "numero_cuota": c.numero_cuota,
                 "monto": float(c.monto), "estado": c.estado,
                 "cliente_nombre": f"{cl.nombre} {cl.apellido}"} for c, p, cl in rows]

    # ── Cobros por mes — últimos 12 meses ──
    mes_i = hoy_d.month - 11
    anio_i = hoy_d.year + (0 if mes_i > 0 else -1)
    if mes_i <= 0:
        mes_i += 12
    inicio_12m = date(anio_i, mes_i, 1)

    cobros_rows = (
        db.query(
            extract("year",  Pago.fecha_pago).label("anio"),
            extract("month", Pago.fecha_pago).label("mes"),
            sqlfunc.sum(Pago.monto_pagado).label("total"),
        )
        .filter(cast(Pago.fecha_pago, SADate) >= inicio_12m)
        .group_by(extract("year", Pago.fecha_pago), extract("month", Pago.fecha_pago))
        .order_by(extract("year", Pago.fecha_pago), extract("month", Pago.fecha_pago))
        .all()
    )
    cobros_por_mes = [{"anio": int(r.anio), "mes": int(r.mes), "total": float(r.total)} for r in cobros_rows]

    mora_result = obtener_cuotas_en_mora(db, limit=100_000)
    return {
        "total_prestado": total_prestado,
        "total_cobrado": total_cobrado,
        "deuda_total": deuda_total,
        "prestamos_activos": prestamos_activos,
        "clientes_con_prestamos": clientes_count,
        "prestamos_por_tipo": prestamos_por_tipo,
        "cuotas_por_estado": cuotas_por_estado,
        "mora": {"total_en_mora": mora_result["total"], "cuotas": mora_result["cuotas"]},
        "filtrado": fecha_desde is not None or fecha_hasta is not None,
        "cobros_hoy": _cuotas_dia(hoy_d),
        "cobros_manana": _cuotas_dia(hoy_d + timedelta(days=1)),
        "cobros_por_mes": cobros_por_mes,
    }


@router.get("/{prestamo_id}/completo")
def detalle_completo(
    prestamo_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    prestamo = (
        db.query(Prestamo)
        .options(
            joinedload(Prestamo.cuotas_rel),
            joinedload(Prestamo.pagos),
            joinedload(Prestamo.cliente),
        )
        .filter(Prestamo.id == prestamo_id)
        .first()
    )
    if not prestamo:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")
    total_pagado = sum(float(p.monto_pagado) for p in prestamo.pagos)
    total_cuotas = sum(float(c.monto) for c in prestamo.cuotas_rel)
    deuda = round(total_cuotas - total_pagado, 2)
    cuotas_sorted = sorted(prestamo.cuotas_rel, key=lambda c: c.numero_cuota)
    pagos_sorted = sorted(prestamo.pagos, key=lambda p: p.fecha_pago, reverse=True)

    hoy = date.today()

    # Compute monto_efectivo
    sum_pagadas  = sum(float(c.monto) for c in cuotas_sorted if c.estado == "pagada")
    sum_parciales = sum(float(c.monto_pagado_parcial or 0) for c in cuotas_sorted if c.estado == "parcial")
    surplus = max(0.0, total_pagado - sum_pagadas - sum_parciales)
    cuotas_data = []
    for c in cuotas_sorted:
        c_monto = float(c.monto)
        monto_parcial = float(c.monto_pagado_parcial or 0)
        # Estado efectivo: cuota pendiente vencida → mostrar como vencida sin tocar la BD
        if c.estado == "pendiente" and c.fecha_vencimiento < hoy:
            estado_efectivo = "vencida"
        else:
            estado_efectivo = c.estado
        if c.estado == "pagada":
            efectivo = 0.0
        elif c.estado == "parcial":
            efectivo = round(c_monto - monto_parcial, 2)
        elif surplus > 0:
            efectivo = max(0.0, round(c_monto - surplus, 2))
            surplus = max(0.0, surplus - c_monto)
        else:
            efectivo = c_monto
        cuotas_data.append({
            "id": c.id, "prestamo_id": c.prestamo_id,
            "numero_cuota": c.numero_cuota,
            "fecha_vencimiento": c.fecha_vencimiento.isoformat(),
            "monto": c_monto, "monto_efectivo": efectivo,
            "monto_pagado_parcial": monto_parcial,
            "estado": estado_efectivo,
        })

    return {
        "prestamo": {
            "id": prestamo.id,
            "cliente_id": prestamo.cliente_id,
            "monto": float(prestamo.monto),
            "interes_total": prestamo.interes_total,
            "cuotas": prestamo.cuotas,
            "monto_cuota": float(prestamo.monto_cuota) if prestamo.monto_cuota else None,
            "fecha_inicio": prestamo.fecha_inicio.isoformat() if prestamo.fecha_inicio else None,
            "estado": prestamo.estado,
            "notas": prestamo.notas or "",
        },
        "cliente": {
            "id": prestamo.cliente.id,
            "nombre": prestamo.cliente.nombre,
            "apellido": prestamo.cliente.apellido,
            "dni": prestamo.cliente.dni,
        } if prestamo.cliente else None,
        "cuotas_rel": cuotas_data,
        "pagos": [
            {
                "id": p.id, "prestamo_id": p.prestamo_id,
                "monto_pagado": float(p.monto_pagado),
                "fecha_pago": p.fecha_pago.isoformat() if p.fecha_pago else None,
                "dias_atraso": p.dias_atraso,
            }
            for p in pagos_sorted
        ],
        "deuda_restante": deuda,
        "total_cuotas": total_cuotas,
        "total_pagado": total_pagado,
    }


@router.post("/{prestamo_id}/cuotas/{cuota_id}/marcar-pagada")
def marcar_cuota_pagada(
    prestamo_id: int,
    cuota_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Marca una cuota individual como pagada y registra el pago."""
    cuota = (
        db.query(Cuota)
        .filter(Cuota.id == cuota_id, Cuota.prestamo_id == prestamo_id)
        .first()
    )
    if not cuota:
        raise HTTPException(status_code=404, detail="Cuota no encontrada")
    if cuota.estado == "pagada":
        raise HTTPException(status_code=400, detail="La cuota ya está pagada")

    # Si hay pago parcial previo, solo se paga el saldo restante
    monto_ya_pagado = float(cuota.monto_pagado_parcial or 0)
    monto_a_pagar = round(float(cuota.monto) - monto_ya_pagado, 2)

    cuota.estado = "pagada"
    db.add(cuota)

    # Registrar pago automático por el saldo restante
    pago = Pago(
        prestamo_id=prestamo_id,
        cuota_id=cuota_id,
        monto_pagado=monto_a_pagar,
        fecha_pago=datetime.now(timezone.utc),
        dias_atraso=max(0, (date.today() - cuota.fecha_vencimiento).days) if cuota.fecha_vencimiento <= date.today() else 0,
    )
    db.add(pago)

    # Verificar si todas las cuotas están pagadas (flush antes del count para incluir el cambio actual)
    db.flush()
    no_pagadas = (
        db.query(Cuota)
        .filter(Cuota.prestamo_id == prestamo_id, Cuota.estado != "pagada")
        .count()
    )
    if no_pagadas == 0:
        prestamo = db.query(Prestamo).filter(Prestamo.id == prestamo_id).first()
        if prestamo:
            prestamo.estado = "finalizado"
            db.add(prestamo)

    db.commit()
    return {"ok": True, "message": f"Cuota #{cuota.numero_cuota} marcada como pagada"}


@router.post("/{prestamo_id}/cuotas/{cuota_id}/desmarcar-pagada")
def desmarcar_cuota_pagada(
    prestamo_id: int,
    cuota_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Revierte el marcado de una cuota como pagada y elimina el pago automático asociado."""
    cuota = (
        db.query(Cuota)
        .filter(Cuota.id == cuota_id, Cuota.prestamo_id == prestamo_id)
        .first()
    )
    if not cuota:
        raise HTTPException(status_code=404, detail="Cuota no encontrada")
    if cuota.estado != "pagada":
        raise HTTPException(status_code=400, detail="La cuota no está marcada como pagada")

    # Restaurar estado: si tenía pago parcial, vuelve a "parcial"; sino, pendiente/vencida
    monto_parcial = float(cuota.monto_pagado_parcial or 0)
    if monto_parcial > 0:
        cuota.estado = "parcial"
    else:
        cuota.estado = "vencida" if cuota.fecha_vencimiento < date.today() else "pendiente"
    db.add(cuota)

    # Eliminar el pago más reciente vinculado a esta cuota (el que completó el pago)
    pago_auto = (
        db.query(Pago)
        .filter(Pago.prestamo_id == prestamo_id, Pago.cuota_id == cuota_id)
        .order_by(Pago.fecha_pago.desc())
        .first()
    )
    if pago_auto:
        db.delete(pago_auto)

    # Si el préstamo estaba finalizado, reactivarlo
    prestamo = db.query(Prestamo).filter(Prestamo.id == prestamo_id).first()
    if prestamo and prestamo.estado == "finalizado":
        prestamo.estado = "activo"
        db.add(prestamo)

    db.commit()
    return {"ok": True, "message": f"Cuota #{cuota.numero_cuota} desmarcada"}


class ParcialPayload(BaseModel):
    monto: float


@router.post("/{prestamo_id}/cuotas/{cuota_id}/pago-parcial")
def pago_parcial_cuota(
    prestamo_id: int,
    cuota_id: int,
    payload: ParcialPayload,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Registra un pago parcial sobre una cuota específica."""
    cuota = (
        db.query(Cuota)
        .filter(Cuota.id == cuota_id, Cuota.prestamo_id == prestamo_id)
        .first()
    )
    if not cuota:
        raise HTTPException(status_code=404, detail="Cuota no encontrada")
    if cuota.estado == "pagada":
        raise HTTPException(status_code=400, detail="La cuota ya está pagada")

    monto_ya_pagado = float(cuota.monto_pagado_parcial or 0)
    monto_restante = round(float(cuota.monto) - monto_ya_pagado, 2)

    if payload.monto <= 0:
        raise HTTPException(status_code=400, detail="El monto debe ser mayor a 0")
    if payload.monto >= monto_restante:
        raise HTTPException(
            status_code=400,
            detail=f"El monto supera o iguala la deuda restante (${monto_restante:,.2f}). Para pago total usá 'Marcar como pagada'.",
        )

    nuevo_parcial = round(monto_ya_pagado + payload.monto, 2)
    cuota.monto_pagado_parcial = nuevo_parcial
    cuota.estado = "parcial"
    db.add(cuota)

    pago = Pago(
        prestamo_id=prestamo_id,
        cuota_id=cuota_id,
        monto_pagado=payload.monto,
        fecha_pago=datetime.now(timezone.utc),
        dias_atraso=max(0, (date.today() - cuota.fecha_vencimiento).days) if cuota.fecha_vencimiento <= date.today() else 0,
    )
    db.add(pago)
    db.commit()
    return {
        "ok": True,
        "monto_pagado_total": nuevo_parcial,
        "monto_restante": round(float(cuota.monto) - nuevo_parcial, 2),
    }


class RefinanciarPayload(BaseModel):
    cuotas_detalle: List[CuotaInput]


@router.post("/{prestamo_id}/refinanciar")
def refinanciar_prestamo(
    prestamo_id: int,
    payload: RefinanciarPayload,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Agrega nuevas cuotas a un préstamo existente (extensión / refinanciación)."""
    prestamo = db.query(Prestamo).filter(Prestamo.id == prestamo_id).first()
    if not prestamo:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    max_num = (
        db.query(sqlfunc.max(Cuota.numero_cuota))
        .filter(Cuota.prestamo_id == prestamo_id)
        .scalar() or 0
    )

    for i, det in enumerate(payload.cuotas_detalle):
        db.add(Cuota(
            prestamo_id=prestamo_id,
            numero_cuota=max_num + i + 1,
            fecha_vencimiento=det.fecha_vencimiento,
            monto=det.monto,
            estado="pendiente",
        ))

    prestamo.cuotas += len(payload.cuotas_detalle)
    if prestamo.estado == "finalizado":
        prestamo.estado = "activo"
    db.add(prestamo)
    db.commit()
    return {"ok": True, "nuevas_cuotas": len(payload.cuotas_detalle), "total_cuotas": prestamo.cuotas}


@router.post("/{prestamo_id}/cancelar")
def cancelar_prestamo(
    prestamo_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Cancela el préstamo: marca todas las cuotas pendientes como pagadas."""
    prestamo = db.query(Prestamo).filter(Prestamo.id == prestamo_id).first()
    if not prestamo:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")
    if prestamo.estado == "finalizado":
        raise HTTPException(status_code=400, detail="El préstamo ya está finalizado")

    cuotas_pendientes = (
        db.query(Cuota)
        .filter(
            Cuota.prestamo_id == prestamo_id,
            Cuota.estado.in_(["pendiente", "vencida", "parcial"]),
        )
        .all()
    )

    total_restante = sum(
        float(c.monto) - float(c.monto_pagado_parcial or 0)
        for c in cuotas_pendientes
    )

    for c in cuotas_pendientes:
        c.estado = "pagada"
        db.add(c)

    # Registrar pago por el total restante
    if total_restante > 0:
        pago = Pago(
            prestamo_id=prestamo_id,
            monto_pagado=total_restante,
            fecha_pago=datetime.now(timezone.utc),
            dias_atraso=0,
        )
        db.add(pago)

    prestamo.estado = "finalizado"
    db.add(prestamo)
    db.commit()

    return {
        "ok": True,
        "message": f"Préstamo #{prestamo_id} cancelado. {len(cuotas_pendientes)} cuotas marcadas como pagadas.",
    }


@router.post("/", response_model=PrestamoRead)
def crear_prestamo(
    payload: PrestamoCreate,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    if len(payload.cuotas_detalle) != payload.cuotas:
        raise HTTPException(
            status_code=400,
            detail=f"Se esperaban {payload.cuotas} cuotas pero se recibieron {len(payload.cuotas_detalle)}",
        )
    prestamo = create_prestamo(
        db,
        cliente_id=payload.cliente_id,
        monto=payload.monto,
        interes_total=payload.interes_total,
        num_cuotas=payload.cuotas,
        cuotas_detalle=payload.cuotas_detalle,
        fecha_inicio=payload.fecha_inicio,
        tipo_prestamo=payload.tipo_prestamo,
    )
    return prestamo


@router.get("/export/xlsx")
def export_prestamos_xlsx(
    estado: Optional[str] = Query(None),
    cliente_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    q = db.query(Prestamo, Cliente).join(Cliente, Prestamo.cliente_id == Cliente.id)
    if estado:
        q = q.filter(Prestamo.estado == estado)
    if cliente_id:
        q = q.filter(Prestamo.cliente_id == cliente_id)
    rows = q.order_by(Prestamo.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Préstamos"

    headers = ["ID", "Cliente", "DNI", "Tipo", "Monto", "Interés (%)", "Cuotas", "Fecha Inicio", "Estado"]
    header_fill = PatternFill("solid", fgColor="0284C7")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for p, c in rows:
        ws.append([
            p.id, f"{c.nombre} {c.apellido}", c.dni,
            p.tipo_prestamo or "mensual",
            float(p.monto), p.interes_total, p.cuotas,
            p.fecha_inicio.isoformat() if p.fecha_inicio else "",
            p.estado,
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prestamos.xlsx"},
    )


@router.get("/")
def listar_prestamos(
    offset: int = 0,
    limit: int = 10,
    estado: Optional[str] = Query(None),
    cliente_id: Optional[int] = Query(None),
    tipo_prestamo: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None, description="id | monto | cliente"),
    sort_desc: bool = Query(True),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    q = (
        db.query(Prestamo)
        .options(joinedload(Prestamo.cliente))
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
    )
    if estado:
        q = q.filter(Prestamo.estado == estado)
    if cliente_id:
        q = q.filter(Prestamo.cliente_id == cliente_id)
    if tipo_prestamo:
        q = q.filter(Prestamo.tipo_prestamo == tipo_prestamo)
    if search:
        term = f"%{search}%"
        q = q.filter(
            or_(
                (Cliente.nombre + " " + Cliente.apellido).ilike(term),
                Cliente.dni.ilike(term),
                cast(Prestamo.id, String).ilike(term),
            )
        )

    if sort_by == "monto":
        order_col = Prestamo.monto.desc() if sort_desc else Prestamo.monto
    elif sort_by == "cliente":
        order_col = (Cliente.apellido.desc() if sort_desc else Cliente.apellido)
    else:  # default: id
        order_col = Prestamo.id.desc() if sort_desc else Prestamo.id
    items = q.order_by(order_col).offset(offset).limit(limit).all()

    # Batch-fetch cuota counts to avoid N+1
    cuotas_stats: dict = {}
    if items:
        prestamo_ids = [p.id for p in items]
        rows = (
            db.query(
                Cuota.prestamo_id,
                sqlfunc.count(Cuota.id).label("total"),
                sqlfunc.sum(case((Cuota.estado == "pagada", 1), else_=0)).label("pagadas"),
            )
            .filter(Cuota.prestamo_id.in_(prestamo_ids))
            .group_by(Cuota.prestamo_id)
            .all()
        )
        cuotas_stats = {r.prestamo_id: {"total": r.total, "pagadas": int(r.pagadas or 0)} for r in rows}

    return [
        {
            "id": p.id,
            "cliente_id": p.cliente_id,
            "cliente_nombre": p.cliente.nombre if p.cliente else "—",
            "cliente_apellido": p.cliente.apellido if p.cliente else "—",
            "cliente_dni": p.cliente.dni if p.cliente else "—",
            "monto": float(p.monto),
            "interes_total": p.interes_total,
            "cuotas": p.cuotas,
            "cuotas_pagadas": cuotas_stats.get(p.id, {}).get("pagadas", 0),
            "cuotas_total": cuotas_stats.get(p.id, {}).get("total", p.cuotas),
            "monto_cuota": float(p.monto_cuota) if p.monto_cuota else None,
            "fecha_inicio": p.fecha_inicio.isoformat() if p.fecha_inicio else None,
            "estado": p.estado,
            "tipo_prestamo": p.tipo_prestamo or "mensual",
        }
        for p in items
    ]


@router.get("/{prestamo_id}", response_model=PrestamoRead)
def obtener_prestamo(
    prestamo_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    p = db.query(Prestamo).filter(Prestamo.id == prestamo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")
    return p


@router.get("/{prestamo_id}/deuda")
def deuda_prestamo(
    prestamo_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    p = db.query(Prestamo).get(prestamo_id)
    if not p:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")
    deuda = calcular_deuda_restante(db, p)
    return {"deuda_restante": deuda}


@router.get("/{prestamo_id}/cuotas", response_model=list[CuotaRead])
def listar_cuotas(
    prestamo_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    p = db.query(Prestamo).get(prestamo_id)
    if not p:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")
    return db.query(Cuota).filter(Cuota.prestamo_id == prestamo_id).order_by(Cuota.numero_cuota).all()


@router.put("/{prestamo_id}/cuotas/{cuota_id}", response_model=CuotaRead)
def actualizar_cuota(
    prestamo_id: int,
    cuota_id: int,
    payload: CuotaUpdate,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    cuota = db.query(Cuota).filter(Cuota.id == cuota_id, Cuota.prestamo_id == prestamo_id).first()
    if not cuota:
        raise HTTPException(status_code=404, detail="Cuota no encontrada")
    if payload.fecha_vencimiento is not None:
        cuota.fecha_vencimiento = payload.fecha_vencimiento
    if payload.monto is not None:
        cuota.monto = payload.monto
    db.add(cuota)
    db.commit()
    db.refresh(cuota)
    return cuota


class NotasPayload(BaseModel):
    notas: str = ""


@router.patch("/{prestamo_id}/notas")
def actualizar_notas(
    prestamo_id: int,
    payload: NotasPayload,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    p = db.query(Prestamo).filter(Prestamo.id == prestamo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")
    p.notas = payload.notas
    db.add(p)
    db.commit()
    return {"ok": True}


@router.delete("/{prestamo_id}")
def eliminar_prestamo(
    prestamo_id: int,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    p = db.query(Prestamo).get(prestamo_id)
    if not p:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")
    db.delete(p)
    db.commit()
    return {"ok": True}
